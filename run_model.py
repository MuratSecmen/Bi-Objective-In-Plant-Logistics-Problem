"""
Internal Logistics MIP — Pickup-Delivery VRP (multi-product, multi-vehicle)

Single-file driver: config + data validator + model builder + solver + reporter.
All numerical parameters come from inputs/config.xlsx. Hard-coded values are
deliberately avoided; the script raises DataConsistencyError on any mismatch
between the model's assumptions and the input data.

Time convention: all model times are in *minutes since shift start*.
Clock-time formatting (HH:MM) in result files re-adds shift_start_clock_min.

Usage
-----
    python run_model.py
    python run_model.py --inputs inputs --config inputs/config.xlsx
    python run_model.py --output-dir results

This is the MIP-only distribution: only the `single_objective` and
`multi_objective` run modes are supported. Solomon, NSGA-II, and the
test-sweep modes have been stripped out.
"""

from __future__ import annotations

import argparse
import logging
import math
import sys
from dataclasses import dataclass, field
from datetime import datetime, time
from itertools import permutations
from pathlib import Path
from typing import Optional

import pandas as pd
import gurobipy as gp
from gurobipy import GRB, quicksum


# =============================================================================
# Logging
# =============================================================================
log = logging.getLogger("internal_logistics")


def configure_logging(level: int = logging.INFO,
                      log_file: Optional[Path] = None,
                      verbose: bool = False) -> None:
    """
    Configure Python logging.

    Default behaviour: write the script's INFO messages ONLY to the log file.
    The console is left clean so Gurobi's own progress output is the only
    thing visible on screen during a solve. Pass verbose=True to also stream
    Python log records to the console with a timestamp prefix.
    """
    handlers: list[logging.Handler] = []
    if log_file is not None:
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))
    if verbose:
        handlers.append(logging.StreamHandler(sys.stdout))
    if not handlers:
        handlers.append(logging.NullHandler())
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=handlers,
        force=True,
    )


# =============================================================================
# Errors
# =============================================================================
class DataConsistencyError(Exception):
    """Raised when input data fails validation against the model's assumptions."""


class ConfigError(Exception):
    """Raised when config.xlsx is malformed or missing required keys."""


# =============================================================================
# Constants
# =============================================================================
OBJECTIVES = {"route_duration", "wait_time"}
RUN_MODES = {"single_objective", "multi_objective"}
OBJ_METHODS = {"augmented_eps", "lexicographic"}

PRODUCT_SET_TO_SHEET = {
    # Legacy fallback mapping. Only consulted if the actual sheet name in
    # products.xlsx doesn't match the product_set_id directly or via its
    # short prefix (e.g. "case1" for "case1_all_distinct").
    "case1_all_distinct": "Sheet1",
    "case2_loc_c": "Sheet2",
    "case3_shared_pickup": "Sheet3",
}

READY_TIME_FORMATS = {"clock", "relative_to_shift"}

REQUIRED_CONFIG_KEYS = {
    "run_mode", "primary_obj", "constraint_obj", "limit_on_constraint_obj",
    "augmentation_weight", "product_set_id", "num_products",
    "shift_start_clock_min", "shift_duration_min",
    "strict_shift_window", "time_limit_seconds", "mip_gap",
    "output_dir", "output_prefix", "write_full_var_sheets",
    "tight_big_M", "break_vehicle_symmetry",
    "add_work_lb_cut", "tight_time_var_bounds",
    "tight_route_activation",
    "use_indicator_constraints",
    "add_product_lb_cut",
    "add_pair_lb_cut", "pair_lb_threshold_min",
    "add_wait_lb_cut",
    "add_reverse_arc_cut", "add_endpoints_cut", "add_adjacency_cut",
    "mtz_type",
    "objective_method",
    "auto_verify", "verify_on_fail", "auto_visualize",
}

VERIFY_FAIL_MODES = {"raise", "warn"}

# Optional parameters: keys allowed in config.xlsx but not required.
# Each "_override" entry defaults to a value computed from the input data.
# Gurobi solver-tuning parameters are intentionally NOT in this list — Gurobi
# uses its own defaults so we never silently impose a tuning choice.
OPTIONAL_CONFIG_DEFAULTS = {
    # Input file names (default: standard names under --inputs dir).
    # Paths are taken relative to --inputs (so don't include the folder).
    "nodes_file":     "nodes.xlsx",
    "vehicles_file":  "vehicles.xlsx",
    "products_file":  "products.xlsx",
    "distances_file": "distances_minutes.xlsx",
    # Ready-time interpretation in products.xlsx.
    #   "clock"             : value is clock time (e.g. "07:10", 430)
    #   "relative_to_shift" : value is minutes after shift start (e.g. 10 = 07:10)
    "ready_time_format": "clock",
    # Per-objective time budget for the lex multi-objective mode. Total
    # TimeLimit (above) still applies to the whole solve; this caps the
    # second stage (secondary objective) so the sweep never burns the
    # full budget on diminishing returns to the second objective.
    "second_obj_time_limit_seconds": 60,
    # Epsilon-constraint sweep step. After each multi-objective solve the
    # constraint on the secondary objective is tightened by this amount
    # (achieved_secondary - eps_step). Smaller -> denser Pareto front,
    # more iterations.
    "eps_step": 1.0,
    # Big-M overrides (default: computed from input data)
    "C_max_minutes_override": None,    # default: max(c_ij)         used in M16
    "e_min_minutes_override": None,    # default: min(e_p_relative)  used in M22
    "Q_max_override": None,            # default: max(q_k)           used in M24, M25
}


# =============================================================================
# Config loader
# =============================================================================
def _truthy(v) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in {"true", "1", "yes", "y", "t"}


def load_config(path: Path) -> dict:
    if not path.exists():
        raise ConfigError(f"config file not found: {path}")
    df = pd.read_excel(path, sheet_name="config")
    cols = {c.lower(): c for c in df.columns}
    if "parameter" not in cols or "value" not in cols:
        raise ConfigError(
            f"config sheet must have columns 'parameter' and 'value' "
            f"(got {df.columns.tolist()})"
        )

    raw = {}
    for _, row in df.iterrows():
        k = row[cols["parameter"]]
        if pd.isna(k):
            continue
        k = str(k).strip()
        if not k or k.startswith("#"):
            continue
        raw[k] = row[cols["value"]]

    cfg: dict = {}
    missing = [k for k in REQUIRED_CONFIG_KEYS
               if k not in raw or pd.isna(raw[k])]
    if missing:
        raise ConfigError(f"missing required config keys: {missing}")

    for k in REQUIRED_CONFIG_KEYS:
        cfg[k] = raw[k]
    for k, default in OPTIONAL_CONFIG_DEFAULTS.items():
        cfg[k] = raw[k] if k in raw and not pd.isna(raw[k]) else default

    # ---- Coercions ----
    cfg["run_mode"] = str(cfg["run_mode"]).strip()
    cfg["primary_obj"] = str(cfg["primary_obj"]).strip()
    cfg["constraint_obj"] = str(cfg["constraint_obj"]).strip()
    cfg["objective_method"] = str(cfg["objective_method"]).strip()
    cfg["product_set_id"] = str(cfg["product_set_id"]).strip()
    cfg["output_dir"] = str(cfg["output_dir"]).strip()
    cfg["output_prefix"] = str(cfg["output_prefix"]).strip()
    cfg["strict_shift_window"] = _truthy(cfg["strict_shift_window"])
    cfg["write_full_var_sheets"] = _truthy(cfg["write_full_var_sheets"])
    cfg["tight_big_M"] = _truthy(cfg["tight_big_M"])
    cfg["break_vehicle_symmetry"] = _truthy(cfg["break_vehicle_symmetry"])
    cfg["add_work_lb_cut"] = _truthy(cfg["add_work_lb_cut"])
    cfg["tight_time_var_bounds"] = _truthy(cfg["tight_time_var_bounds"])
    cfg["tight_route_activation"] = _truthy(cfg["tight_route_activation"])
    cfg["use_indicator_constraints"] = _truthy(cfg["use_indicator_constraints"])
    cfg["add_product_lb_cut"] = _truthy(cfg["add_product_lb_cut"])
    cfg["add_pair_lb_cut"] = _truthy(cfg["add_pair_lb_cut"])
    cfg["pair_lb_threshold_min"] = float(cfg["pair_lb_threshold_min"])
    cfg["add_wait_lb_cut"] = _truthy(cfg["add_wait_lb_cut"])
    cfg["add_reverse_arc_cut"] = _truthy(cfg["add_reverse_arc_cut"])
    cfg["add_endpoints_cut"] = _truthy(cfg["add_endpoints_cut"])
    cfg["add_adjacency_cut"] = _truthy(cfg["add_adjacency_cut"])
    cfg["mtz_type"] = str(cfg["mtz_type"]).strip().lower()
    if cfg["mtz_type"] == "ss_lifted":
        raise ConfigError(
            "mtz_type='ss_lifted' has been removed: the coefficients "
            "produced an invalid lifting that cut off feasible route "
            "endpoints (u_first = 1, u_last = U). Use 'dl_lifted' "
            "(Desrochers-Laporte), which is the strongest known "
            "polynomial single-constraint lift of MTZ."
        )
    if cfg["mtz_type"] not in {"base", "dl_lifted"}:
        raise ConfigError(
            f"mtz_type must be one of base, dl_lifted "
            f"(got {cfg['mtz_type']!r})"
        )
    cfg["auto_verify"] = _truthy(cfg["auto_verify"])
    cfg["auto_visualize"] = _truthy(cfg["auto_visualize"])
    cfg["verify_on_fail"] = str(cfg["verify_on_fail"]).strip().lower()

    # ---- Input file names (strip whitespace; tolerate Excel quirks) ----
    for fk in ("nodes_file", "vehicles_file", "products_file", "distances_file"):
        cfg[fk] = str(cfg[fk]).strip()

    # ---- Ready-time format ----
    cfg["ready_time_format"] = str(cfg["ready_time_format"]).strip().lower()
    if cfg["ready_time_format"] not in READY_TIME_FORMATS:
        raise ConfigError(
            f"ready_time_format must be one of {sorted(READY_TIME_FORMATS)}, "
            f"got {cfg['ready_time_format']!r}"
        )

    cfg["second_obj_time_limit_seconds"] = int(cfg["second_obj_time_limit_seconds"])
    if cfg["second_obj_time_limit_seconds"] < 1:
        raise ConfigError(
            "second_obj_time_limit_seconds must be >= 1 "
            f"(got {cfg['second_obj_time_limit_seconds']})"
        )

    cfg["eps_step"] = float(cfg["eps_step"])
    if cfg["eps_step"] <= 0:
        raise ConfigError(
            f"eps_step must be > 0 (got {cfg['eps_step']})"
        )

    np_val = cfg["num_products"]
    cfg["num_products"] = (
        None if str(np_val).strip().lower() in {"all", "none", ""}
        else int(np_val)
    )

    for k in ("shift_start_clock_min",
              "shift_duration_min", "time_limit_seconds"):
        cfg[k] = int(cfg[k])
    for k in ("limit_on_constraint_obj", "augmentation_weight", "mip_gap"):
        cfg[k] = float(cfg[k])
    for k in ("C_max_minutes_override", "e_min_minutes_override",
              "Q_max_override"):
        cfg[k] = float(cfg[k]) if cfg[k] is not None and not pd.isna(cfg[k]) else None

    # ---- Validations ----
    if cfg["run_mode"] not in RUN_MODES:
        raise ConfigError(f"run_mode must be in {RUN_MODES}, got '{cfg['run_mode']}'")
    if cfg["primary_obj"] not in OBJECTIVES:
        raise ConfigError(f"primary_obj must be in {OBJECTIVES}, got '{cfg['primary_obj']}'")
    if cfg["constraint_obj"] not in OBJECTIVES:
        raise ConfigError(f"constraint_obj must be in {OBJECTIVES}, got '{cfg['constraint_obj']}'")
    if cfg["primary_obj"] == cfg["constraint_obj"]:
        raise ConfigError("primary_obj and constraint_obj must be different")
    if cfg["objective_method"] not in OBJ_METHODS:
        raise ConfigError(
            f"objective_method must be in {OBJ_METHODS}, "
            f"got '{cfg['objective_method']}'"
        )
    if cfg["verify_on_fail"] not in VERIFY_FAIL_MODES:
        raise ConfigError(
            f"verify_on_fail must be in {VERIFY_FAIL_MODES}, "
            f"got '{cfg['verify_on_fail']}'"
        )
    # product_set_id is no longer validated against PRODUCT_SET_TO_SHEET.
    # The sheet lookup at load_instance time is permissive and accepts
    # exact name, short prefix (case1), and the legacy mapping.
    # default_max_routes is no longer a config parameter — per-vehicle
    # max_route values are read from vehicles.xlsx.
    if cfg["limit_on_constraint_obj"] < 0:
        raise ConfigError("limit_on_constraint_obj must be >= 0")
    if cfg["shift_duration_min"] <= 0:
        raise ConfigError("shift_duration_min must be > 0")
    if cfg["augmentation_weight"] < 0:
        raise ConfigError("augmentation_weight must be >= 0")

    return cfg


# =============================================================================
# Time helpers
# =============================================================================
def ready_to_clock_min(v) -> int:
    """Convert ready_time entry to clock minutes from midnight (no shift offset)."""
    if pd.isna(v):
        raise DataConsistencyError("ready_time is NaN")
    if isinstance(v, (pd.Timestamp, datetime)):
        return int(v.hour) * 60 + int(v.minute)
    if isinstance(v, time):
        return int(v.hour) * 60 + int(v.minute)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return int(v)
    s = str(v).strip()
    dt = pd.to_datetime(s, errors="coerce")
    if pd.notna(dt):
        return int(dt.hour) * 60 + int(dt.minute)
    if ":" in s:
        hh, mm = s.split(":")[:2]
        return int(hh) * 60 + int(mm)
    return int(float(s))


def minutes_to_hhmm(minutes) -> str:
    if minutes is None or pd.isna(minutes):
        return ""
    minutes = float(minutes)
    h = int(minutes // 60)
    m = int(round(minutes - 60 * h))
    return f"{h:02d}:{m:02d}"


# =============================================================================
# Instance
# =============================================================================
@dataclass
class Instance:
    N: list           # incl. depot 'h'
    Nw: list          # work-centers (N \ {h})
    K: list           # vehicle ids
    R: list           # 1..max(max_routes[k]) — global ceiling across fleet
    P: list           # product ids
    c: dict           # (i,j) -> minutes
    e: dict           # p -> shift-relative ready time
    s_load: dict
    s_unload: dict
    q_p: dict
    o: dict
    d: dict
    q_k: dict
    max_routes: dict        # vehicle_id -> max route count (from vehicles.xlsx)
    T_max: float
    C_max: float
    e_min: float
    Q_max: float
    M16: float          # value used in constraint (16); depends on tight_big_M
    M20: float          # value used in constraint (20)
    M22_p: dict         # PRODUCT-INDEXED Big-M used in constraint (22)
    M24: float
    M25: float
    M16_pdf: float      # PDF nominal (T_max + C_max), kept for reporting
    M22_pdf: float      # PDF nominal (T_max - e_min), kept for reporting
    big_m_mode: str     # "tight" or "pdf"
    vehicle_symmetry_groups: list   # list of lists of vehicle ids that are identical
    min_complete: dict              # p -> min single-product route completion time
    pair_lb_cuts: list              # list of (p, q, min_pair_complete) above threshold
    U: int
    config: dict

    @property
    def shift_offset(self) -> int:
        return self.config["shift_start_clock_min"]

    def routes_of(self, k) -> list:
        """Route indices available to vehicle k: [1, 2, ..., max_routes[k]]."""
        return list(range(1, int(self.max_routes[k]) + 1))

    def last_route_of(self, k) -> int:
        """The last (highest-indexed) route of vehicle k."""
        return int(self.max_routes[k])

    @property
    def KR_pairs(self) -> list:
        """All valid (vehicle, route) pairs across the fleet."""
        return [(k, r) for k in self.K for r in self.routes_of(k)]


# =============================================================================
# Helpers for min-completion-time computation (pair / singleton cuts)
# =============================================================================
def _min_route_through_nodes(perm, node_unloads, node_loads, node_ready, c):
    """Simulate a route h -> perm -> h, returning total elapsed minutes.

    At each node: first do unloads (sum_unload), then wait until any pickup's
    ready time, then loads (sum_load). Captures constraints (17)–(19).
    """
    time = 0.0
    prev = "h"
    for node in perm:
        time += c[(prev, node)]
        ts = max(time + node_unloads.get(node, 0.0),
                 node_ready.get(node, 0.0))
        time = ts + node_loads.get(node, 0.0)
        prev = node
    time += c[(prev, "h")]
    return time


def _min_complete_one(o, d, e, sl, su, c):
    """Minimum route completion time for one product alone."""
    if o == d:
        return _min_route_through_nodes(
            [o], {o: su}, {o: sl}, {o: e}, c
        )
    return _min_route_through_nodes(
        [o, d], {d: su}, {o: sl}, {o: e}, c
    )


def _min_complete_pair(o1, d1, e1, sl1, su1,
                        o2, d2, e2, sl2, su2, c):
    """Minimum route completion time for a pair (1, 2) on the same route.

    Handles all node-coincidence cases (shared origin, shared destination,
    origin = destination across products, etc.) by enumerating permutations
    of the UNIQUE work-centre node set.
    """
    unique_nodes = list({o1, d1, o2, d2})
    node_unloads = {n: 0.0 for n in unique_nodes}
    node_loads = {n: 0.0 for n in unique_nodes}
    node_ready = {n: 0.0 for n in unique_nodes}

    # Aggregate events at each visited node.
    node_loads[o1] += sl1
    node_loads[o2] += sl2
    node_unloads[d1] += su1
    node_unloads[d2] += su2
    node_ready[o1] = max(node_ready[o1], e1)
    node_ready[o2] = max(node_ready[o2], e2)

    best = float("inf")
    for perm in permutations(unique_nodes):
        pos = {n: i for i, n in enumerate(perm)}
        # Pickup-before-delivery per product (trivial when o == d).
        if o1 != d1 and pos[o1] > pos[d1]:
            continue
        if o2 != d2 and pos[o2] > pos[d2]:
            continue
        t = _min_route_through_nodes(
            perm, node_unloads, node_loads, node_ready, c
        )
        if t < best:
            best = t
    return best


# =============================================================================
# Data loader (with strict validation — no silent skips)
# =============================================================================
def resolve_products_sheet(xlsx_path: Path, product_set_id: str) -> str:
    """Locate the sheet in the products workbook that matches `product_set_id`.

    Tries, in order:
      1. exact match against the product_set_id (e.g. 'case1_all_distinct')
      2. short prefix before the first underscore ('case1')
      3. the legacy PRODUCT_SET_TO_SHEET mapping ('Sheet1')
      4. case-insensitive variants of (1) and (2)

    Raises DataConsistencyError listing the available sheets if nothing
    matches. This is what lets the same product_set_id work whether the
    student names sheets 'Sheet1', 'case1', or 'case1_all_distinct'.
    """
    xl = pd.ExcelFile(xlsx_path)
    available = list(xl.sheet_names)

    # 1. Exact
    if product_set_id in available:
        return product_set_id
    # 2. Short prefix
    short = product_set_id.split("_", 1)[0] if "_" in product_set_id else product_set_id
    if short != product_set_id and short in available:
        return short
    # 3. Legacy mapping
    legacy = PRODUCT_SET_TO_SHEET.get(product_set_id)
    if legacy and legacy in available:
        return legacy
    # 4. Case-insensitive matching
    avail_lower = {s.lower(): s for s in available}
    for cand in (product_set_id, short):
        if cand.lower() in avail_lower:
            return avail_lower[cand.lower()]

    raise DataConsistencyError(
        f"could not find a sheet for product_set_id={product_set_id!r} in "
        f"{xlsx_path.name}. Available sheets: {available}"
    )


def load_instance(inputs_dir: Path, cfg: dict) -> Instance:
    log.info("Loading inputs from %s", inputs_dir)

    nodes_path     = inputs_dir / cfg["nodes_file"]
    vehicles_path  = inputs_dir / cfg["vehicles_file"]
    products_path  = inputs_dir / cfg["products_file"]
    distances_path = inputs_dir / cfg["distances_file"]
    log.info("Input files: nodes=%s, vehicles=%s, products=%s, distances=%s",
             cfg["nodes_file"], cfg["vehicles_file"],
             cfg["products_file"], cfg["distances_file"])

    # ---- Nodes ----
    if not nodes_path.exists():
        raise DataConsistencyError(f"nodes file not found: {nodes_path}")
    nodes_df = pd.read_excel(nodes_path)
    if "node_id" not in nodes_df.columns:
        raise DataConsistencyError(
            f"{cfg['nodes_file']} must contain column 'node_id'"
        )
    nodes_df["node_id"] = nodes_df["node_id"].astype(str).str.strip()
    if (nodes_df["node_id"] == "").any() or nodes_df["node_id"].isna().any():
        raise DataConsistencyError(
            f"{cfg['nodes_file']} contains empty node_id rows"
        )
    if nodes_df["node_id"].duplicated().any():
        dups = nodes_df.loc[nodes_df["node_id"].duplicated(), "node_id"].tolist()
        raise DataConsistencyError(
            f"duplicate node_id(s) in {cfg['nodes_file']}: {dups}"
        )
    if "h" not in nodes_df["node_id"].tolist():
        raise DataConsistencyError(
            f"{cfg['nodes_file']} must contain depot node 'h'"
        )
    N = nodes_df["node_id"].tolist()
    Nw = [n for n in N if n != "h"]

    # ---- Vehicles ----
    if not vehicles_path.exists():
        raise DataConsistencyError(f"vehicles file not found: {vehicles_path}")
    veh_df = pd.read_excel(vehicles_path)
    required = {"vehicle_id", "capacity_m2", "max_route", "active"}
    if not required.issubset(set(veh_df.columns)):
        raise DataConsistencyError(
            f"{cfg['vehicles_file']} must have columns {sorted(required)} "
            f"(got {veh_df.columns.tolist()})"
        )
    veh_df["vehicle_id"] = veh_df["vehicle_id"].astype(str).str.strip()
    if veh_df["vehicle_id"].duplicated().any():
        dups = veh_df.loc[veh_df["vehicle_id"].duplicated(), "vehicle_id"].tolist()
        raise DataConsistencyError(f"duplicate vehicle_id(s): {dups}")
    if (veh_df["capacity_m2"].astype(float) <= 0).any():
        raise DataConsistencyError("all vehicle capacity_m2 must be > 0")

    # active: per-vehicle on/off flag. Accepts 1/0, true/false, yes/no.
    # Inactive vehicles are dropped from the model entirely: they do not
    # appear in K, no variables/constraints reference them, no symmetry
    # group includes them. At least one vehicle must be active.
    if veh_df["active"].isna().any():
        bad = veh_df.loc[veh_df["active"].isna(), "vehicle_id"].tolist()
        raise DataConsistencyError(
            f"{cfg['vehicles_file']} column 'active' has NaN value(s) "
            f"for vehicle(s): {bad}"
        )
    active_mask = veh_df["active"].apply(_truthy).tolist()
    inactive_ids = [v for v, a in zip(veh_df["vehicle_id"], active_mask) if not a]
    if inactive_ids:
        log.info("Inactive vehicles (excluded from model): %s", inactive_ids)
        print(f"[load] Inactive vehicles excluded: {inactive_ids}")
    veh_df = veh_df.loc[active_mask].reset_index(drop=True)
    if veh_df.empty:
        raise DataConsistencyError(
            f"all vehicles in {cfg['vehicles_file']} are marked inactive — "
            "the model has no fleet to dispatch"
        )

    # max_route: positive integer per active vehicle.
    max_route_raw = veh_df["max_route"]
    if max_route_raw.isna().any():
        raise DataConsistencyError(
            f"{cfg['vehicles_file']} column 'max_route' has NaN values"
        )
    try:
        max_route_int = max_route_raw.astype(int)
    except Exception as exc:
        raise DataConsistencyError(
            f"{cfg['vehicles_file']} column 'max_route' must be integer "
            f"(got: {max_route_raw.tolist()}): {exc}"
        )
    if (max_route_int < 1).any():
        bad = [(vid, mr) for vid, mr in zip(veh_df["vehicle_id"], max_route_int)
               if mr < 1]
        raise DataConsistencyError(
            f"max_route must be >= 1 for every active vehicle; offenders: {bad}"
        )
    K = veh_df["vehicle_id"].tolist()
    q_k = dict(zip(K, veh_df["capacity_m2"].astype(float)))
    max_routes = dict(zip(K, max_route_int.astype(int).tolist()))
    R = list(range(1, max(max_routes.values()) + 1))
    log.info("Active vehicles (%d): %s", len(K), K)
    log.info("Per-vehicle max_route: %s; global R = 1..%d",
             max_routes, R[-1])

    # ---- Detect identical-vehicle groups for symmetry breaking ----
    # Two vehicles are "identical" iff every column other than vehicle_id has
    # the same value. We use ALL non-id columns so that adding a new column
    # later (e.g. vehicle_type, fixed_cost) automatically prevents wrongly
    # pairing vehicles that differ on it.
    sym_cols = [c for c in veh_df.columns if c != "vehicle_id"]
    veh_df["_attr_signature"] = veh_df[sym_cols].apply(
        lambda row: tuple(row.tolist()), axis=1
    )
    vehicle_symmetry_groups = []
    for sig, grp in veh_df.groupby("_attr_signature", sort=False):
        ids = grp["vehicle_id"].tolist()
        if len(ids) > 1:
            vehicle_symmetry_groups.append(ids)
    log.info("Vehicle symmetry groups (by columns %s): %s",
             sym_cols, vehicle_symmetry_groups or "none")

    # ---- Products ----
    if not products_path.exists():
        raise DataConsistencyError(f"products file not found: {products_path}")
    sheet = resolve_products_sheet(products_path, cfg["product_set_id"])
    products_df = pd.read_excel(products_path, sheet_name=sheet)
    log.info("%s::%s loaded with %d rows  (product_set_id=%s)",
             cfg["products_file"], sheet, len(products_df),
             cfg["product_set_id"])

    required_p = {"product_id", "origin", "destination", "ready_time",
                  "load_time", "unload_time", "area_m2"}
    if not required_p.issubset(set(products_df.columns)):
        raise DataConsistencyError(
            f"products.xlsx::{sheet} must have columns {sorted(required_p)} "
            f"(got {products_df.columns.tolist()})"
        )

    if cfg["num_products"] is not None:
        n = int(cfg["num_products"])
        if n > len(products_df):
            raise DataConsistencyError(
                f"num_products={n} but sheet has only {len(products_df)} rows"
            )
        products_df = products_df.head(n).copy()
        log.info("Sliced first %d products (num_products=%d)", len(products_df), n)

    for col in ("product_id", "origin", "destination"):
        products_df[col] = products_df[col].astype(str).str.strip()

    if products_df["product_id"].duplicated().any():
        dups = products_df.loc[products_df["product_id"].duplicated(), "product_id"].tolist()
        raise DataConsistencyError(f"duplicate product_id(s): {dups}")

    bad_o = sorted(set(products_df["origin"]) - set(N))
    if bad_o:
        raise DataConsistencyError(
            f"product origins not in nodes.xlsx: {bad_o}"
        )
    bad_d = sorted(set(products_df["destination"]) - set(N))
    if bad_d:
        raise DataConsistencyError(
            f"product destinations not in nodes.xlsx: {bad_d}"
        )

    for col in ("load_time", "unload_time", "area_m2"):
        if products_df[col].isna().any():
            raise DataConsistencyError(f"products column '{col}' has NaN values")
        if (products_df[col].astype(float) < 0).any():
            raise DataConsistencyError(f"products column '{col}' has negative values")

    P = products_df["product_id"].tolist()

    # Interpret ready_time according to ready_time_format:
    #   "clock"             : value is clock time of day. Numeric like 430,
    #                         string "07:10", or a timestamp/time object —
    #                         all converted to clock minutes from midnight,
    #                         then made shift-relative by subtracting
    #                         shift_start_clock_min.
    #   "relative_to_shift" : value is already minutes-after-shift-start.
    #                         No subtraction performed. Strings like "10"
    #                         and numbers like 10 both accepted.
    fmt = cfg["ready_time_format"]
    if fmt == "clock":
        e_clock = {p: ready_to_clock_min(v)
                   for p, v in zip(P, products_df["ready_time"])}
        e_relative = {p: e_clock[p] - cfg["shift_start_clock_min"] for p in P}
    elif fmt == "relative_to_shift":
        def _to_rel(v):
            if pd.isna(v):
                raise DataConsistencyError("ready_time is NaN")
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return int(v)
            # Reject clock-style values explicitly: datetime / time objects
            # and strings containing ':' are almost certainly clock format,
            # not minutes-after-shift-start.
            if isinstance(v, (pd.Timestamp, datetime, time)):
                raise DataConsistencyError(
                    f"ready_time value {v!r} looks like clock time but "
                    f"ready_time_format='relative_to_shift'. "
                    f"Set ready_time_format='clock' in config.xlsx, "
                    f"or change the column to integer minutes-after-shift-start."
                )
            s = str(v).strip()
            if ":" in s or "-" in s:
                raise DataConsistencyError(
                    f"ready_time value {s!r} looks like clock time but "
                    f"ready_time_format='relative_to_shift'. "
                    f"Set ready_time_format='clock' in config.xlsx, "
                    f"or change the column to integer minutes-after-shift-start."
                )
            try:
                return int(float(s))
            except ValueError as exc:
                raise DataConsistencyError(
                    f"could not parse ready_time {s!r} as an integer "
                    f"(minutes after shift start): {exc}"
                ) from exc
        e_relative = {p: _to_rel(v)
                      for p, v in zip(P, products_df["ready_time"])}
    else:
        raise ConfigError(f"unknown ready_time_format: {fmt!r}")

    out_of_window = {p: e_relative[p]
                     for p in P
                     if e_relative[p] < 0
                     or e_relative[p] > cfg["shift_duration_min"]}
    if out_of_window:
        msg = (f"products outside shift window [0, {cfg['shift_duration_min']}] "
               f"min after shift start: {out_of_window}")
        if cfg["strict_shift_window"]:
            raise DataConsistencyError(msg)
        log.warning(msg)

    s_load = dict(zip(P, products_df["load_time"].astype(float)))
    s_unload = dict(zip(P, products_df["unload_time"].astype(float)))
    q_p = dict(zip(P, products_df["area_m2"].astype(float)))
    o = dict(zip(P, products_df["origin"]))
    d = dict(zip(P, products_df["destination"]))

    # ---- Restrict N / Nw to nodes referenced by the selected products ----
    # The nodes file may declare more work stations than the current product
    # subset actually uses. Carrying unused nodes inflates the model, the
    # result.xlsx (per-node sheets), and the visualisations. We restrict the
    # active node set to {h} ∪ {o_p, d_p : p ∈ P}, preserving the original
    # order from nodes.xlsx so downstream displays stay stable.
    used_nodes = {"h"} | set(o.values()) | set(d.values())
    full_N = list(N)
    N = [n for n in full_N if n in used_nodes]
    Nw = [n for n in N if n != "h"]
    unused = [n for n in full_N if n not in used_nodes]
    msg = (
        f"Active nodes: {len(N)} of {len(full_N)} "
        f"(skipping {len(unused)} unused work stations"
        + (f": {', '.join(unused)}" if 0 < len(unused) <= 12
           else f": {', '.join(unused[:12])}, ..." if unused
           else "")
        + ")"
    )
    log.info(msg)
    print(f"[load] {msg}")

    # ---- Distances ----
    # Primary path comes from cfg["distances_file"]; a Turkish-named legacy
    # file is also tolerated when the user hasn't overridden the config.
    dist_path = distances_path
    if not dist_path.exists():
        legacy = inputs_dir / "distances - dakika.xlsx"
        if legacy.exists():
            log.warning("distance file %s not found; using legacy %s",
                        cfg["distances_file"], legacy.name)
            dist_path = legacy
        else:
            raise DataConsistencyError(
                f"distances file not found: {dist_path} "
                f"(also tried legacy 'distances - dakika.xlsx')"
            )
    dist_df = pd.read_excel(dist_path)
    if not {"from_node", "to_node", "duration_min"}.issubset(set(dist_df.columns)):
        raise DataConsistencyError(
            f"{dist_path.name} must have columns from_node, to_node, duration_min"
        )
    dist_df["from_node"] = dist_df["from_node"].astype(str).str.strip()
    dist_df["to_node"] = dist_df["to_node"].astype(str).str.strip()
    dist_df["duration_min"] = pd.to_numeric(dist_df["duration_min"], errors="coerce")
    if dist_df["duration_min"].isna().any():
        raise DataConsistencyError("distances has non-numeric duration_min")
    if (dist_df["duration_min"] < 0).any():
        raise DataConsistencyError("distances has negative duration_min")

    dist_df = dist_df[
        (dist_df["from_node"].isin(N))
        & (dist_df["to_node"].isin(N))
        & (dist_df["from_node"] != dist_df["to_node"])
    ]
    if dist_df.duplicated(subset=["from_node", "to_node"]).any():
        raise DataConsistencyError("distances has duplicate (from,to) rows")
    have = set(zip(dist_df["from_node"], dist_df["to_node"]))
    need = {(i, j) for i in N for j in N if i != j}
    missing = need - have
    if missing:
        raise DataConsistencyError(
            f"distances missing {len(missing)} pair(s); "
            f"first 10: {sorted(missing)[:10]}"
        )
    c = {(r["from_node"], r["to_node"]): float(r["duration_min"])
         for _, r in dist_df.iterrows()}

    # ---- Derived parameters ----
    T_max = float(cfg["shift_duration_min"])
    C_max = (float(cfg["C_max_minutes_override"])
             if cfg["C_max_minutes_override"] is not None else max(c.values()))
    e_min = (float(cfg["e_min_minutes_override"])
             if cfg["e_min_minutes_override"] is not None else min(e_relative.values()))
    Q_max = (float(cfg["Q_max_override"])
             if cfg["Q_max_override"] is not None else max(q_k.values()))

    # ---- Big-M coefficients ----
    # Shift-relative time means T_start = 0 in the model.
    # PDF nominal values (eq. 37-41):
    M16_pdf = T_max + C_max          # eq. 37
    M22_pdf = T_max - e_min          # eq. 39 (uniform across products)
    # Tight values (provably valid; argument in MODEL_CODE_AUDIT.md):
    #   M16 = T_max          (the +C_max term in (16) is x-dependent and vanishes when x=0)
    #   M22(p) = T_max + s_unload[p] - e_p   (per-product, replaces the uniform value)
    M16_tight = T_max
    M22_p_tight = {p: T_max + s_unload[p] - e_relative[p] for p in P}

    # Pick the active set based on the config flag.
    big_m_mode = "tight" if cfg["tight_big_M"] else "pdf"
    if cfg["tight_big_M"]:
        M16 = M16_tight
        M22_p = M22_p_tight
    else:
        M16 = M16_pdf
        M22_p = {p: M22_pdf for p in P}

    M20 = T_max          # eq. 38, no tightening proposed in this round
    M24 = Q_max          # eq. 40
    M25 = Q_max          # eq. 41
    U = len(Nw)

    log.info("Sets: |N|=%d, |Nw|=%d, |K|=%d, |R|=%d, |P|=%d",
             len(N), len(Nw), len(K), len(R), len(P))
    log.info("Derived: T_max=%.0f, C_max=%.2f, e_min=%.2f, Q_max=%.2f, U=%d",
             T_max, C_max, e_min, Q_max, U)
    log.info("Big-M mode: %s", big_m_mode)
    log.info("  M16(pdf)=%.2f  M16(used)=%.2f", M16_pdf, M16)
    log.info("  M22(pdf,uniform)=%.2f", M22_pdf)
    log.info("  M22(used) min=%.2f, max=%.2f, mean=%.2f",
             min(M22_p.values()), max(M22_p.values()),
             sum(M22_p.values()) / len(M22_p))
    log.info("  M20=%.2f  M24=%.2f  M25=%.2f", M20, M24, M25)

    # ---- Self-loop product check (o_p == d_p disallowed) ----
    self_loops = [p for p in P if o[p] == d[p]]
    if self_loops:
        raise DataConsistencyError(
            f"products with origin == destination are not allowed: {self_loops}"
        )

    # ---- Per-product min completion time (used by product_lb_cut) ----
    min_complete = {
        p: _min_complete_one(o[p], d[p], e_relative[p],
                              s_load[p], s_unload[p], c)
        for p in P
    }

    # ---- Pre-compute pair_lb cuts (filtered by threshold) ----
    # delta(p, q) = min_pair_complete(p, q) - max(min_complete(p), min_complete(q))
    # Emit a cut only when delta > pair_lb_threshold_min.
    pair_lb_cuts: list = []
    pair_lb_threshold = float(cfg["pair_lb_threshold_min"])
    if cfg["add_pair_lb_cut"]:
        bucket_counts = {"<=0": 0, "(0, 2]": 0, "(2, 5]": 0,
                          "(5, 10]": 0, "> 10": 0}
        P_list = list(P)
        for i in range(len(P_list)):
            for j in range(i + 1, len(P_list)):
                p1, p2 = P_list[i], P_list[j]
                mp = _min_complete_pair(
                    o[p1], d[p1], e_relative[p1], s_load[p1], s_unload[p1],
                    o[p2], d[p2], e_relative[p2], s_load[p2], s_unload[p2],
                    c,
                )
                delta = mp - max(min_complete[p1], min_complete[p2])
                # Bucket for histogram
                if delta <= 0:
                    bucket_counts["<=0"] += 1
                elif delta <= 2:
                    bucket_counts["(0, 2]"] += 1
                elif delta <= 5:
                    bucket_counts["(2, 5]"] += 1
                elif delta <= 10:
                    bucket_counts["(5, 10]"] += 1
                else:
                    bucket_counts["> 10"] += 1
                if delta > pair_lb_threshold:
                    pair_lb_cuts.append((p1, p2, mp))

        log.info("Pair-LB delta histogram (threshold = %.2f):",
                 pair_lb_threshold)
        for label, count in bucket_counts.items():
            log.info("  delta %s : %d pairs", label, count)
        log.info("Pair-LB will emit %d pairs (filtered).",
                 len(pair_lb_cuts))

    return Instance(
        N=N, Nw=Nw, K=K, R=R, P=P,
        c=c, e=e_relative, s_load=s_load, s_unload=s_unload,
        q_p=q_p, o=o, d=d, q_k=q_k, max_routes=max_routes,
        T_max=T_max, C_max=C_max, e_min=e_min, Q_max=Q_max,
        M16=M16, M20=M20, M22_p=M22_p, M24=M24, M25=M25,
        M16_pdf=M16_pdf, M22_pdf=M22_pdf,
        big_m_mode=big_m_mode,
        vehicle_symmetry_groups=vehicle_symmetry_groups,
        min_complete=min_complete,
        pair_lb_cuts=pair_lb_cuts,
        U=U,
        config=cfg,
    )


# =============================================================================
# Model builder
# =============================================================================
def build_model(inst: Instance, *, primary: str, constraint: str,
                limit: float, weight: float, method: str):
    """Build the PD-VRP MIP per the PDF model (constraints 1-36).

    primary, constraint ∈ {'route_duration', 'wait_time'}.
    method ∈ {'augmented_eps', 'lexicographic'}.

    augmented_eps: single-objective `primary_expr + weight * constraint_expr`.
    lexicographic: Gurobi multi-objective with priority(primary) > priority(secondary);
                    the augmentation_weight is unused in this branch.

    The ε-constraint `constraint_expr ≤ limit` is added in BOTH methods so
    multi-objective ε-sweep is consistent across the two.
    """
    if primary == constraint:
        raise ValueError("primary_obj and constraint_obj must differ")
    if method not in OBJ_METHODS:
        raise ValueError(f"method must be in {OBJ_METHODS}, got '{method}'")

    m = gp.Model(f"InternalLogistics_{primary}_primary")

    # ---- Decision variables (per-vehicle route index sets) ----
    # Each vehicle k has its own route range 1..max_routes[k] from
    # vehicles.xlsx. Variables are created only for valid (k, r) pairs,
    # so a vehicle with max_route = 1 contributes one route's worth of
    # variables while a vehicle with max_route = 3 contributes three.
    # Tight upper bounds on time variables (recommendation B). Under the
    # shift-window assumption (already implicit in the PDF Big-Ms), every
    # feasible time variable lies in [0, T_max]. Adding ub=T_max is free
    # tightening of the LP polyhedron and does not cut any feasible point.
    time_ub = inst.T_max if inst.config["tight_time_var_bounds"] else GRB.INFINITY

    KR = inst.KR_pairs  # list of (k, r) pairs across the whole fleet

    arc_keys = [(i, j, k, r) for (k, r) in KR
                for i in inst.N for j in inst.N if i != j]
    f_keys     = [(p, k, r) for (k, r) in KR for p in inst.P]
    y_keys     = [(j, k, r) for (k, r) in KR for j in inst.N]
    ta_keys    = [(j, k, r) for (k, r) in KR for j in inst.N]
    ts_keys    = [(j, k, r) for (k, r) in KR for j in inst.Nw]
    u_keys     = [(j, k, r) for (k, r) in KR for j in inst.Nw]
    delta_keys = [(j, k, r) for (k, r) in KR for j in inst.Nw]

    x     = m.addVars(arc_keys, vtype=GRB.BINARY, name="x")
    f     = m.addVars(f_keys,   vtype=GRB.BINARY, name="f")
    w     = m.addVars(inst.P,   vtype=GRB.CONTINUOUS, lb=0.0, name="w")
    y     = m.addVars(y_keys,   vtype=GRB.CONTINUOUS, lb=0.0, name="y")
    ta    = m.addVars(ta_keys,  vtype=GRB.CONTINUOUS, lb=0.0,
                      ub=time_ub, name="ta")
    td    = m.addVars(ta_keys,  vtype=GRB.CONTINUOUS, lb=0.0,
                      ub=time_ub, name="td")
    ts    = m.addVars(ts_keys,  vtype=GRB.CONTINUOUS, lb=0.0,
                      ub=time_ub, name="ts")
    u     = m.addVars(u_keys,   vtype=GRB.INTEGER, lb=0, ub=inst.U, name="u")
    delta = m.addVars(delta_keys, vtype=GRB.CONTINUOUS,
                      lb=-GRB.INFINITY, name="delta")
    log.info("Per-vehicle route counts: %s; total |KR|=%d",
             {k: inst.max_routes[k] for k in inst.K}, len(KR))

    # Shift-relative model: T_start = 0, so route_duration is the sum over
    # vehicles of the arrival time at the depot in EACH vehicle's last route.
    # (Each vehicle's "last route" is its max_routes[k], which can differ.)
    route_duration_expr = quicksum(ta["h", k, inst.last_route_of(k)]
                                   for k in inst.K)
    total_wait_expr = quicksum(w[p] for p in inst.P)
    obj_exprs = {"route_duration": route_duration_expr,
                 "wait_time": total_wait_expr}

    # ---- (1)/(2) Objective: augmented or lexicographic ----
    if method == "augmented_eps":
        # Single objective: primary + 10^-3 · secondary (PDF formulation)
        m.setObjective(obj_exprs[primary] + weight * obj_exprs[constraint],
                       GRB.MINIMIZE)
    else:  # method == "lexicographic"
        # Two objectives, priority(primary) > priority(secondary).
        # Gurobi solves primary to optimality first, then minimizes secondary
        # subject to primary being held at its optimum.
        #
        # abstol=0 and reltol=0 force STRICT lex ordering: the primary may not
        # worsen by even a numerical tolerance during secondary optimization.
        # Defaults are 1e-6 / 0.0; we set both to zero so each Pareto point
        # we report is on the true Pareto frontier with no slack.
        m.ModelSense = GRB.MINIMIZE
        m.NumObj = 2
        m.setObjectiveN(obj_exprs[primary],    index=0, priority=2,
                        abstol=0.0, reltol=0.0,
                        name=f"min_{primary}")
        m.setObjectiveN(obj_exprs[constraint], index=1, priority=1,
                        abstol=0.0, reltol=0.0,
                        name=f"min_{constraint}")

    # ---- (3) Epsilon constraint ----
    m.addConstr(obj_exprs[constraint] <= limit, name=f"c3_eps_on_{constraint}")

    # ---- (4)-(9) Route structure ----
    for k in inst.K:
        for r in inst.routes_of(k):
            out_h = quicksum(x["h", j, k, r] for j in inst.Nw)
            in_h = quicksum(x[j, "h", k, r] for j in inst.Nw)
            m.addConstr(out_h == in_h, name=f"c4[{k},{r}]")
            m.addConstr(out_h <= 1, name=f"c5[{k},{r}]")
            # Constraint (6) — route activation. Two forms:
            #   - PDF (loose):   Sum x_ijkr <= (2|P| + 1) * Sum f_pkr
            #   - tight:         Sum x_ijkr <= 2 * Sum f_pkr + Sum x_h,j,k,r
            # Both kill the route when Sum f = 0 (Case A in the audit) and both
            # cap the arc count when the route is used. The tight form is
            # `2 * (#products on this route) + 1`, where the +1 piggybacks on
            # constraint (5) so it only "kicks in" when a depot departure is
            # made. The tight form is strictly stronger than the PDF form.
            total_arcs = quicksum(x[i, j, k, r] for i in inst.N
                                  for j in inst.N if i != j)
            total_f = quicksum(f[p, k, r] for p in inst.P)
            if inst.config["tight_route_activation"]:
                m.addConstr(
                    total_arcs <= 2 * total_f + out_h,
                    name=f"c6[{k},{r}]",
                )
            else:
                m.addConstr(
                    total_arcs <= (2 * len(inst.P) + 1) * total_f,
                    name=f"c6[{k},{r}]",
                )
    for j in inst.Nw:
        for k in inst.K:
            for r in inst.routes_of(k):
                inflow = quicksum(x[i, j, k, r] for i in inst.N if i != j)
                outflow = quicksum(x[j, i, k, r] for i in inst.N if i != j)
                relevant = quicksum(
                    f[p, k, r] for p in inst.P
                    if inst.o[p] == j or inst.d[p] == j
                )
                m.addConstr(inflow <= relevant, name=f"c7[{j},{k},{r}]")
                m.addConstr(inflow == outflow, name=f"c8[{j},{k},{r}]")
                m.addConstr(inflow <= 1, name=f"c9[{j},{k},{r}]")

    # ---- (10)-(12) Product assignment + visit pickup/delivery ----
    for p in inst.P:
        m.addConstr(
            quicksum(f[p, k, r] for (k, r) in inst.KR_pairs) == 1,
            name=f"c10[{p}]",
        )
        op, dp = inst.o[p], inst.d[p]
        for k in inst.K:
            for r in inst.routes_of(k):
                m.addConstr(
                    quicksum(x[i, op, k, r] for i in inst.N if i != op)
                    >= f[p, k, r],
                    name=f"c11[{p},{k},{r}]",
                )
                m.addConstr(
                    quicksum(x[i, dp, k, r] for i in inst.N if i != dp)
                    >= f[p, k, r],
                    name=f"c12[{p},{k},{r}]",
                )

    # ---- Vehicle symmetry breaking (valid only across truly identical vehicles) ----
    # When two or more vehicles share the same capacity (and thus the same role
    # in the model), the solver wastes effort exploring assignments that differ
    # only by the label of the vehicle. For each maximal group of identical
    # vehicles, we enforce a "load monotonicity" constraint on consecutive
    # pairs: the earlier-indexed vehicle must carry at least as many products
    # (over the whole shift) as the next one. This collapses the symmetry orbit
    # while still allowing every workload partition that was reachable before.
    # Valid only across vehicles that are truly interchangeable (same q_k,
    # same depot, no time-windowed differences), which is why we restrict it
    # to pre-computed `vehicle_symmetry_groups`.
    if inst.config["break_vehicle_symmetry"]:
        for group in inst.vehicle_symmetry_groups:
            for k_a, k_b in zip(group, group[1:]):
                m.addConstr(
                    quicksum(f[p, k_a, r] for p in inst.P for r in inst.routes_of(k_a))
                    >= quicksum(f[p, k_b, r] for p in inst.P for r in inst.routes_of(k_b)),
                    name=f"sym[{k_a}>={k_b}]",
                )

    # ---- Work-LB cut (recommendation A): per-vehicle aggregate valid inequality ----
    # Derivation: route duration ≥ travel + service (since wait ≥ 0). The chain
    # via constraints (14)+(21) gives ta[h,k,|R|] ≥ Σ_r duration(r). Together:
    #
    #   ta[h, k, |R|] ≥ Σ_{i,j,r} c_ij · x_ijkr + Σ_{p,r} (s^l_p + s^u_p) · f_pkr
    #
    # No integer-feasible solution is cut; LP root bound jumps from 0 to a
    # finite positive value because the routing constraints (10), (11), (12)
    # force x and f to accumulate.
    if inst.config["add_work_lb_cut"]:
        for k in inst.K:
            travel_expr = quicksum(
                inst.c[(i, j)] * x[i, j, k, r]
                for i in inst.N for j in inst.N
                if i != j
                for r in inst.routes_of(k)
            )
            service_expr = quicksum(
                (inst.s_load[p] + inst.s_unload[p]) * f[p, k, r]
                for p in inst.P for r in inst.routes_of(k)
            )
            m.addConstr(
                ta["h", k, inst.last_route_of(k)] >= travel_expr + service_expr,
                name=f"work_lb[{k}]",
            )

    # ---- Per-product completion-time lower bound (pairwise (depot, p) cut) ----
    # For every (product p, vehicle k), if p is assigned to k on any route,
    # the final depot arrival of k is at least:
    #
    #     min_complete(p) = max(c(h, o_p), e_p)
    #                       + s^l_p + c(o_p, d_p) + s^u_p + c(d_p, h)
    #
    # Derived by chaining (16)–(20) under triangle-inequality on c. The
    # constraint scales by Σ_r f_{p,k,r}, which equals 1 iff p is on k.
    # No integer-feasible solution is cut.
    #
    # EMPIRICAL NOTE (2026-05-11): On the case-1 15-product instance, enabling
    # this cut tightens the root LP bound noticeably but adds 45 constraints
    # (|P| · |K|) that interact poorly with Gurobi's primal heuristics and
    # automatic cut generation. Net effect on that instance: solver finishes
    # in ~60 s without the cut and stalls at ~21% gap with the cut.
    # The cut is therefore OFF by default; toggle it on for experiments on
    # larger instances (20+ products) where add_work_lb_cut is too loose.
    if inst.config["add_product_lb_cut"]:
        for p in inst.P:
            for k in inst.K:
                last_route_idx = inst.last_route_of(k)
                m.addConstr(
                    ta["h", k, last_route_idx]
                    >= inst.min_complete[p]
                       * quicksum(f[p, k, r] for r in inst.routes_of(k)),
                    name=f"product_lb[{p},{k}]",
                )

    # ---- Pair lower-bound cut (2-product subset version) ----
    # For each pair (p1, p2) where min_pair_complete(p1, p2) is strictly larger
    # than max(min_complete(p1), min_complete(p2)) by more than the threshold,
    # and for each vehicle k that can fit both:
    #
    #   ta[h, k, |R|] >= min_pair_complete(p1, p2) *
    #                    ( Σ_r f_{p1,k,r} + Σ_r f_{p2,k,r} - 1 )
    #
    # When both products are on k, the RHS multiplier is 1 and the bound binds.
    # When only one is on k, the multiplier is 0 and the cut is trivial.
    # When neither is on k, the multiplier is -1 and the cut is trivial.
    if inst.config["add_pair_lb_cut"] and inst.pair_lb_cuts:
        emitted = 0
        for (p1, p2, mp_val) in inst.pair_lb_cuts:
            q_pair = inst.q_p[p1] + inst.q_p[p2]
            for k in inst.K:
                if q_pair > inst.q_k[k]:
                    continue   # cannot share a route, skip
                last_route_idx = inst.last_route_of(k)
                m.addConstr(
                    ta["h", k, last_route_idx]
                    >= mp_val * (
                        quicksum(f[p1, k, r] for r in inst.routes_of(k))
                        + quicksum(f[p2, k, r] for r in inst.routes_of(k))
                        - 1
                    ),
                    name=f"pair_lb[{p1},{p2},{k}]",
                )
                emitted += 1
        log.info("Pair-LB cuts emitted: %d", emitted)

    # ---- Per-product wait lower-bound cut ----
    # Derivation: in any feasible solution shipping p on (k, r),
    #
    #   t^s_{o_p,k,r} >= e_p           (constraint 18)
    #   t^d_{o_p,k,r} >= t^s + s^l_p   (constraint 19)
    #   t^a_{d_p,k,r} >= t^d_{o_p,k,r} + (actual travel time o_p -> d_p)
    #
    # The truck may visit intermediate nodes between o_p and d_p (serving
    # other products), so the "actual travel time" along its path is the
    # sum of arc costs c_{ij} along the SEQUENCE used — NOT necessarily
    # c(o_p, d_p) directly. A safe lower bound is the **shortest path**
    # from o_p to d_p in the c-graph: any path the truck might take has
    # total travel time >= sp(o_p, d_p). If c satisfies the triangle
    # inequality, sp(o_p, d_p) == c(o_p, d_p) and the cut is exact;
    # otherwise sp can be strictly smaller.
    #
    # Combined with constraint (22):
    #
    #      w_p  >=  s^l_p + sp(o_p, d_p) + s^u_p     for every product p.
    #
    # This bound is independent of (k, r), uses no big-M, and provides a
    # strong constant lower bound on the wait-time objective at the root.
    if inst.config["add_wait_lb_cut"]:
        # Floyd-Warshall over the existing arc matrix. |N|^3 = trivial at
        # our scale. Use only when the cut is actually requested.
        node_list = list(inst.N)
        n_idx = {n: i for i, n in enumerate(node_list)}
        n_nodes = len(node_list)
        INF = float("inf")
        sp = [[INF] * n_nodes for _ in range(n_nodes)]
        for i in range(n_nodes):
            sp[i][i] = 0.0
        for (a, b), val in inst.c.items():
            sp[n_idx[a]][n_idx[b]] = float(val)
        for k_idx in range(n_nodes):
            row_k = sp[k_idx]
            for i in range(n_nodes):
                dik = sp[i][k_idx]
                if dik == INF:
                    continue
                row_i = sp[i]
                for j in range(n_nodes):
                    via = dik + row_k[j]
                    if via < row_i[j]:
                        row_i[j] = via

        wait_lb_total = 0.0
        relaxed = 0
        for p in inst.P:
            op_i, dp_i = n_idx[inst.o[p]], n_idx[inst.d[p]]
            travel_lb = sp[op_i][dp_i]
            direct = inst.c.get((inst.o[p], inst.d[p]), travel_lb)
            if travel_lb + 1e-9 < direct:
                relaxed += 1
            lb_p = inst.s_load[p] + travel_lb + inst.s_unload[p]
            m.addConstr(w[p] >= lb_p, name=f"wait_lb[{p}]")
            wait_lb_total += lb_p
        log.info("Wait-LB cuts emitted: %d (Σ w_p >= %.2f min); "
                 "%d cuts loosened by triangle-inequality violations",
                 len(inst.P), wait_lb_total, relaxed)

    # =========================================================================
    # Precedence-strengthened cuts for pickup-and-delivery routes
    # -------------------------------------------------------------------------
    # Each product p imposes the precedence o_p < d_p on the route that ships
    # it (constraint (32) on visit-rank, constraint (20) on time). Both are
    # correct at integer-feasible solutions but LP-weak because of their
    # respective big-M coefficients (U for (32), T_max for (20)).
    #
    # The three cuts below live directly on x and f (no big-M), and target
    # specific fractional patterns the LP likes to produce. All three are
    # valid in every integer-feasible solution (they only forbid arrangements
    # that already violate pickup-before-delivery), so no optimum is removed.
    # =========================================================================

    # ---- Reverse-arc cut ----
    # If product p is shipped on (k, r), the truck cannot traverse the direct
    # arc d_p -> o_p on the same route — that would visit the destination
    # before the origin. Encoded with a sum-<= form on x and f:
    #
    #     x_{d_p, o_p, k, r}  +  f_{p, k, r}  <=  1
    #
    # When f = 1 the cut forces x = 0; when f = 0 the cut is satisfied
    # trivially (x in {0, 1}). At the LP, fractional weight on the reverse
    # arc gets clipped by however much f leans toward 1.
    if inst.config["add_reverse_arc_cut"]:
        emitted = 0
        for p in inst.P:
            op, dp = inst.o[p], inst.d[p]
            if op in inst.Nw and dp in inst.Nw and op != dp:
                for k in inst.K:
                    for r in inst.routes_of(k):
                        m.addConstr(
                            x[dp, op, k, r] + f[p, k, r] <= 1,
                            name=f"rev_arc[{p},{k},{r}]",
                        )
                        emitted += 1
        log.info("Reverse-arc cuts emitted: %d", emitted)

    # ---- Forbidden-endpoints cut ----
    # If product p is shipped on (k, r), the truck cannot have d_p as the
    # first work station visited (right after the depot) nor o_p as the last
    # work station visited (right before returning to the depot). Either
    # would violate pickup-before-delivery.
    #
    #     x_{h, d_p, k, r}  +  f_{p, k, r}  <=  1     (d_p not first)
    #     x_{o_p, h, k, r}  +  f_{p, k, r}  <=  1     (o_p not last)
    #
    # Attacks fractional LP patterns where the depot fractionally heads
    # straight to a destination (typically cheap, hence LP-attractive).
    if inst.config["add_endpoints_cut"]:
        emitted = 0
        for p in inst.P:
            op, dp = inst.o[p], inst.d[p]
            if op not in inst.Nw or dp not in inst.Nw:
                continue
            for k in inst.K:
                for r in inst.routes_of(k):
                    m.addConstr(
                        x["h", dp, k, r] + f[p, k, r] <= 1,
                        name=f"endpt_first[{p},{k},{r}]",
                    )
                    m.addConstr(
                        x[op, "h", k, r] + f[p, k, r] <= 1,
                        name=f"endpt_last[{p},{k},{r}]",
                    )
                    emitted += 2
        log.info("Endpoints cuts emitted: %d", emitted)

    # ---- 1-vertex adjacency cut (simplest tournament constraint) ----
    # If product p is shipped on (k, r), the truck cannot do the 1-vertex
    # detour d_p -> v -> o_p for any intermediate v in N_w \ {o_p, d_p}.
    # That sequence would visit the destination, then a single other
    # station, then the origin — still a precedence violation.
    #
    #     x_{d_p, v, k, r}  +  x_{v, o_p, k, r}  +  f_{p, k, r}  <=  2
    #
    # Implied by (32) at integer level (chained MTZ gives u_{o_p} = u_{d_p}+2,
    # contradicting (32)'s u_{d_p} >= u_{o_p} + 1) but lives on x and f with
    # small coefficients, so it cuts a different slice of the LP polyhedron.
    # On PD-VRPs this is typically the highest-yield precedence cut.
    #
    # Cost: |P| * (|N_w| - 2) * |K| * |R|. Polynomial; cheap enough to add
    # statically. A 2-vertex extension would scale quadratically in |N_w|
    # and would be a candidate for lazy separation instead.
    if inst.config["add_adjacency_cut"]:
        emitted = 0
        for p in inst.P:
            op, dp = inst.o[p], inst.d[p]
            if op not in inst.Nw or dp not in inst.Nw:
                continue
            for v in inst.Nw:
                if v == op or v == dp:
                    continue
                for k in inst.K:
                    for r in inst.routes_of(k):
                        m.addConstr(
                            x[dp, v, k, r] + x[v, op, k, r] + f[p, k, r] <= 2,
                            name=f"adj1[{p},{v},{k},{r}]",
                        )
                        emitted += 1
        log.info("1-vertex adjacency cuts emitted: %d", emitted)

    # ---- (13)-(15) Route timing ----
    for k in inst.K:
        m.addConstr(td["h", k, 1] == 0, name=f"c13[{k}]")
    for k in inst.K:
        for r in inst.routes_of(k)[1:]:
            m.addConstr(td["h", k, r] >= ta["h", k, r - 1], name=f"c14[{k},{r}]")
            m.addConstr(ta["h", k, r] >= ta["h", k, r - 1], name=f"c15[{k},{r}]")

    # ---- (16) Time consistency on traversed arcs ----
    use_ind = inst.config["use_indicator_constraints"]
    for (i, j, k, r) in arc_keys:
        if use_ind:
            m.addGenConstrIndicator(
                x[i, j, k, r], True,
                ta[j, k, r] - td[i, k, r] >= inst.c[(i, j)],
                name=f"c16[{i},{j},{k},{r}]",
            )
        else:
            m.addConstr(
                ta[j, k, r]
                >= td[i, k, r] + inst.c[(i, j)] * x[i, j, k, r]
                   - inst.M16 * (1 - x[i, j, k, r]),
                name=f"c16[{i},{j},{k},{r}]",
            )

    # ---- (17)-(19) Service times ----
    for j in inst.Nw:
        for k in inst.K:
            for r in inst.routes_of(k):
                unload = quicksum(inst.s_unload[p] * f[p, k, r]
                                  for p in inst.P if inst.d[p] == j)
                load = quicksum(inst.s_load[p] * f[p, k, r]
                                for p in inst.P if inst.o[p] == j)
                m.addConstr(ts[j, k, r] >= ta[j, k, r] + unload,
                            name=f"c17[{j},{k},{r}]")
                m.addConstr(td[j, k, r] >= ts[j, k, r] + load,
                            name=f"c19[{j},{k},{r}]")
    for p in inst.P:
        op = inst.o[p]
        if op == "h":
            continue  # ts is defined only on Nw
        for k in inst.K:
            for r in inst.routes_of(k):
                m.addConstr(
                    ts[op, k, r] >= inst.e[p] * f[p, k, r],
                    name=f"c18[{p},{k},{r}]",
                )

    # ---- (20)-(22) Pickup/delivery precedence, depot timing, waiting ----
    for p in inst.P:
        op, dp = inst.o[p], inst.d[p]
        for k in inst.K:
            for r in inst.routes_of(k):
                if use_ind:
                    m.addGenConstrIndicator(
                        f[p, k, r], True,
                        ta[dp, k, r] - td[op, k, r] >= 0,
                        name=f"c20[{p},{k},{r}]",
                    )
                else:
                    m.addConstr(
                        ta[dp, k, r] >= td[op, k, r] - inst.M20 * (1 - f[p, k, r]),
                        name=f"c20[{p},{k},{r}]",
                    )
    for k in inst.K:
        for r in inst.routes_of(k):
            m.addConstr(ta["h", k, r] >= td["h", k, r], name=f"c21[{k},{r}]")
    for p in inst.P:
        dp = inst.d[p]
        ep = inst.e[p]
        sp = inst.s_unload[p]
        m22_p = inst.M22_p[p]   # per-product Big-M (constraint 22)
        for k in inst.K:
            for r in inst.routes_of(k):
                if use_ind:
                    m.addGenConstrIndicator(
                        f[p, k, r], True,
                        w[p] - ta[dp, k, r] >= sp - ep,
                        name=f"c22[{p},{k},{r}]",
                    )
                else:
                    m.addConstr(
                        w[p] >= ta[dp, k, r] + sp - ep
                                - m22_p * (1 - f[p, k, r]),
                        name=f"c22[{p},{k},{r}]",
                    )

    # ---- (23)-(27) Load flow ----
    for j in inst.Nw:
        for k in inst.K:
            for r in inst.routes_of(k):
                load_in = quicksum(inst.q_p[p] * f[p, k, r]
                                   for p in inst.P if inst.o[p] == j)
                load_out = quicksum(inst.q_p[p] * f[p, k, r]
                                    for p in inst.P if inst.d[p] == j)
                # PDF eq. (23): Delta_jkr >= Sum_{o_p=j} f q_p - Sum_{d_p=j} f q_p.
                # We TIGHTEN to equality because the variable's stated meaning
                # is "net change in load at node j". With the PDF's >= form,
                # Delta becomes free to be larger than the net change (it has
                # no objective coefficient and no upper bound from elsewhere),
                # which makes y also drift above the true physical load and
                # breaks the semantics of all post-solve interpretation. The
                # equality form preserves every feasible (x, f, time, route)
                # combination and the optimum's objective is unchanged.
                m.addConstr(delta[j, k, r] == load_in - load_out,
                            name=f"c23[{j},{k},{r}]")
                m.addConstr(y[j, k, r] <= inst.q_k[k],
                            name=f"c26[{j},{k},{r}]")
    for i in inst.Nw:
        for j in inst.Nw:
            if i == j:
                continue
            for k in inst.K:
                for r in inst.routes_of(k):
                    if use_ind:
                        # (24) + (25) collapse to a single equality when x = 1
                        m.addGenConstrIndicator(
                            x[i, j, k, r], True,
                            y[j, k, r] - y[i, k, r] - delta[j, k, r] == 0,
                            name=f"c24_25[{i},{j},{k},{r}]",
                        )
                    else:
                        m.addConstr(
                            y[j, k, r] >= y[i, k, r] + delta[j, k, r]
                                           - inst.M24 * (1 - x[i, j, k, r]),
                            name=f"c24[{i},{j},{k},{r}]",
                        )
                        m.addConstr(
                            y[j, k, r] <= y[i, k, r] + delta[j, k, r]
                                           + inst.M25 * (1 - x[i, j, k, r]),
                            name=f"c25[{i},{j},{k},{r}]",
                        )
    for k in inst.K:
        for r in inst.routes_of(k):
            m.addConstr(y["h", k, r] == 0, name=f"c27[{k},{r}]")

    # ---- (28) Route monotonicity ----
    for k in inst.K:
        for r in inst.routes_of(k)[:-1]:
            m.addConstr(
                quicksum(x["h", j, k, r] for j in inst.Nw)
                >= quicksum(x["h", j, k, r + 1] for j in inst.Nw),
                name=f"c28[{k},{r}]",
            )

    # ---- (29)-(32) MTZ + visit-order pickup precedence ----
    # The u_{j,k,r} variable encodes the position (visit rank) of work-station
    # node j on vehicle k's route r. By convention u = 0 if j is not visited
    # on (k, r) and u in {1, ..., U} otherwise, where U = |N_w| (set during
    # data loading). Constraints (29)-(31) calibrate u to that interpretation;
    # constraint (32) imposes pickup-before-delivery on the visit order.

    # ---- (29) Subtour elimination via Miller-Tucker-Zemlin ----
    # If arc i -> j is traversed on (k, r), the rank strictly increases by at
    # least 1. Cycles among work stations would force u_i >= u_i + 1, which is
    # impossible, so (29) eliminates subtours. Two formulations are offered
    # via the `mtz_type` config flag; both define the same integer-feasible
    # set, but their LP relaxations differ.
    #
    #   "base"      -- Miller-Tucker-Zemlin (1960):
    #                  u_j >= u_i + 1 - U(1 - x_ij)
    #                  Big-M of size U switches the constraint OFF when the
    #                  arc is unused. Weaker LP relaxation.
    #
    #   "dl_lifted" -- Desrochers-Laporte (1991):
    #                  u_j >= u_i + 1 - U(1 - x_ij) + (U - 2) x_ji
    #                  Adds a lifting term that uses the reverse arc. Validity:
    #                  if x_ji = 1 in an integer solution, MTZ on the reverse
    #                  arc already implies u_j <= u_i - 1, so the lifting bakes
    #                  that consequence into the LP polyhedron without removing
    #                  any integer point. Strictly tighter LP than "base", and
    #                  is the strongest known polynomial single-constraint lift
    #                  of MTZ: any coefficient larger than (U-2) on x_ji can be
    #                  violated by the integer-feasible configuration where
    #                  x_ji = 1 forces u_i - u_j = 1.
    #
    # An earlier "ss_lifted" option (Sarin-Sherali-style) was removed: its
    # coefficients (U-1)*x_ij + (U-3)*x_ji and RHS (U-2) over-tighten the
    # right-hand side and forbid valid integer-feasible configurations such
    # as the route endpoints (u_first = 1, u_last = U). The Sarin-Sherali
    # paper achieves strength by introducing auxiliary variables, not by a
    # single-cut coefficient strengthening.
    mtz_type = inst.config["mtz_type"]
    for i in inst.Nw:
        for j in inst.Nw:
            if i == j:
                continue
            for k in inst.K:
                for r in inst.routes_of(k):
                    xij = x[i, j, k, r]
                    xji = x[j, i, k, r]
                    if mtz_type == "base":
                        rhs = u[i, k, r] + 1 - inst.U * (1 - xij)
                    else:  # dl_lifted
                        rhs = (u[i, k, r] + 1
                               - inst.U * (1 - xij)
                               + (inst.U - 2) * xji)
                    m.addConstr(u[j, k, r] >= rhs,
                                name=f"c29[{i},{j},{k},{r}]")

    # ---- (30)-(31) "u calibration" inequalities ----
    # Together they pin u to the visit-rank interpretation:
    #   (30) u_{jkr} <= U * indeg(j)  -- if no arc enters j, indeg = 0 and the
    #        rank is forced to 0 (so unvisited work stations have rank 0).
    #   (31) u_{jkr} >= indeg(j)      -- if some arc enters j, indeg = 1 (by
    #        the flow constraints), so the visited node's rank is >= 1.
    # Without these two, u could drift freely for unvisited nodes and the LP
    # relaxation of (29) and (32) would be much weaker.
    for j in inst.Nw:
        for k in inst.K:
            for r in inst.routes_of(k):
                indeg = quicksum(x[i, j, k, r] for i in inst.N if i != j)
                m.addConstr(u[j, k, r] <= inst.U * indeg,
                            name=f"c30[{j},{k},{r}]")
                m.addConstr(u[j, k, r] >= indeg,
                            name=f"c31[{j},{k},{r}]")

    # ---- (32) Pickup-before-delivery precedence on visit order ----
    # If product p is shipped on (k, r), the destination's rank must exceed
    # the origin's rank by at least 1. The big-M term -U(1 - f_{pkr}) turns
    # the constraint OFF when the product is not assigned to (k, r). This is
    # the rank-scale companion to time-scale constraint (20); both encode the
    # same precedence, but (32) uses a much smaller big-M (U = |N_w|) and is
    # therefore the LP-stronger statement.
    for p in inst.P:
        op, dp = inst.o[p], inst.d[p]
        if op in inst.Nw and dp in inst.Nw:
            for k in inst.K:
                for r in inst.routes_of(k):
                    m.addConstr(
                        u[dp, k, r] >= u[op, k, r] + 1
                                        - inst.U * (1 - f[p, k, r]),
                        name=f"c32[{p},{k},{r}]",
                    )

    vars_ = {"x": x, "f": f, "w": w, "y": y, "ta": ta, "td": td,
             "ts": ts, "u": u, "delta": delta,
             "route_duration_expr": route_duration_expr,
             "total_wait_expr": total_wait_expr}
    return m, vars_


# =============================================================================
# Solve
# =============================================================================
def solve_model(m: gp.Model, *, cfg: dict, gurobi_log: Path,
                compute_iis_on_infeasible: bool = True) -> None:
    """Set only the user-facing solver parameters (time limit, gap) and
    the log file. Everything else is left at Gurobi's defaults.

    Parameters
    ----------
    compute_iis_on_infeasible : bool, default True
        If the model returns INFEASIBLE, run an IIS computation and write
        the ILP file. Set to False for the multi-objective sweep's
        tightening iterations, where infeasibility is the *expected*
        termination signal and IIS noise would only clutter the log.
    """
    m.setParam("TimeLimit", cfg["time_limit_seconds"])
    m.setParam("MIPGap", cfg["mip_gap"])
    m.setParam("LogFile", str(gurobi_log))
    # m.Params.MIPFocus = 3  # focus on improving the bound
    # m.Params.Cuts = 3  # aggressive cut generation
    # m.Params.Symmetry = 2  # detect symmetry beyond what we've broken statically
    # m.Params.VarBranch = 3  # strong branching (slower per node but tighter)

    # Multi-objective (lex) mode: cap the SECONDARY objective's solve time.
    # Each objective in a multi-obj model has its own parameter environment
    # accessible via getMultiobjEnv(idx); setting TimeLimit there bounds the
    # time spent on that particular stage. We leave the primary stage
    # (index 0) unbounded so it can fully resolve the objective-1 optimum,
    # then cap stage 2 (index 1, the secondary objective) to avoid burning
    # the rest of the global TimeLimit on diminishing returns.
    if getattr(m, "NumObj", 1) > 1:
        try:
            second_env = m.getMultiobjEnv(1)
            second_env.setParam(
                "TimeLimit",
                int(cfg["second_obj_time_limit_seconds"]),
            )
        except gp.GurobiError as exc:
            log.warning("Could not set per-objective time limit: %s", exc)

    m.update()
    m.optimize()

    # ---- On infeasibility, compute and report the IIS ----
    if m.Status == GRB.INFEASIBLE and not compute_iis_on_infeasible:
        log.info("Model is INFEASIBLE; IIS computation skipped "
                 "(compute_iis_on_infeasible=False).")
        return  # caller is responsible for handling the infeasibility
    # Gurobi's irreducible inconsistent subsystem (IIS) is the smallest set of
    # constraints + variable bounds that, taken together, are still infeasible.
    # Dropping any one of them would make the IIS feasible. Reading the IIS
    # constraint list directly tells you which model constraint(s) are in
    # conflict — far more useful than just "Model is infeasible".
    if m.Status == GRB.INFEASIBLE:
        try:
            log.warning(
                "Model reported INFEASIBLE — computing IIS for diagnosis..."
            )
            print("[solve] Model is INFEASIBLE — computing IIS...")
            m.computeIIS()
            ilp_path = Path(str(gurobi_log)).with_suffix(".ilp")
            m.write(str(ilp_path))

            iis_constrs = [c.constrName for c in m.getConstrs() if c.IISConstr]
            iis_var_lb  = [v.varName for v in m.getVars() if v.IISLB]
            iis_var_ub  = [v.varName for v in m.getVars() if v.IISUB]

            log.info("IIS written to %s", ilp_path)
            log.info("IIS constraint count: %d", len(iis_constrs))
            log.info("IIS lb-bound count:   %d", len(iis_var_lb))
            log.info("IIS ub-bound count:   %d", len(iis_var_ub))

            # Group constraints by their family (the prefix before [) so the
            # report is readable even when 200+ are in conflict.
            from collections import Counter
            family = Counter(
                c.split("[", 1)[0] for c in iis_constrs
            )
            print(f"[iis] {len(iis_constrs)} constraints in conflict:")
            for fam, n in family.most_common():
                print(f"      {fam:<30s} {n}")
            if iis_var_lb:
                print(f"[iis] {len(iis_var_lb)} variable lower bounds in IIS:")
                for v in iis_var_lb[:20]:
                    print(f"      {v}")
                if len(iis_var_lb) > 20:
                    print(f"      ... and {len(iis_var_lb) - 20} more")
            if iis_var_ub:
                print(f"[iis] {len(iis_var_ub)} variable upper bounds in IIS:")
                for v in iis_var_ub[:20]:
                    print(f"      {v}")
                if len(iis_var_ub) > 20:
                    print(f"      ... and {len(iis_var_ub) - 20} more")
            print(f"[iis] Full ILP file written to {ilp_path}")
            print(f"[iis] Open it in a text editor — every constraint there "
                  f"is part of the infeasibility.")
        except gp.GurobiError as exc:
            # IIS computation can fail in multi-obj or with some attributes;
            # log but don't re-raise so the run still terminates cleanly.
            log.error("IIS computation failed: %s", exc)
            print(f"[iis] IIS computation failed: {exc}")


# =============================================================================
# Reporter
# =============================================================================
def write_results(m: gp.Model, vars_: dict, inst: Instance,
                  *, output_path: Path) -> Optional[dict]:
    if m.SolCount == 0:
        log.error("No feasible solution; status=%s", m.status)
        return None

    x = vars_["x"]; f = vars_["f"]; w = vars_["w"]; y = vars_["y"]
    ta = vars_["ta"]; td = vars_["td"]; ts = vars_["ts"]
    u = vars_["u"]; delta = vars_["delta"]
    route_duration_val = sum(ta["h", k, inst.last_route_of(k)].X for k in inst.K)
    total_wait_val = sum(w[p].X for p in inst.P)
    off = inst.shift_offset

    def stamp(v):
        return minutes_to_hhmm(v + off) if v is not None and not pd.isna(v) else ""

    # Per-objective bounds and gap (Gurobi exposes ObjBound / MIPGap as scalars
    # only in single-objective mode; in multi-objective mode they must be
    # queried via the ObjNumber parameter). Wrap every read in try/except so
    # both modes work without diverging the reporter.
    def _safe_attr(model, name, default=float("nan")):
        try:
            return getattr(model, name)
        except Exception:
            return default

    is_multi_obj = (getattr(m, "NumObj", 1) > 1)
    if is_multi_obj:
        # In lex mode use the primary objective's value as obj_value
        # (the one Gurobi minimized first, with priority 2).
        if inst.config["primary_obj"] == "route_duration":
            obj_value = route_duration_val
        else:
            obj_value = total_wait_val
        # Per-objective bounds via ObjNumber are not portable; report NaN here
        # and rely on the route_duration_min / total_wait_min columns instead.
        best_bound = float("nan")
        mip_gap = float("nan")
    else:
        obj_value = _safe_attr(m, "ObjVal")
        best_bound = _safe_attr(m, "ObjBound")
        mip_gap = _safe_attr(m, "MIPGap")

    summary = {
        "product_set_id": inst.config["product_set_id"],
        "num_products": len(inst.P),
        "num_vehicles": len(inst.K),
        "num_routes_per_vehicle": "; ".join(
            f"{k}:{inst.last_route_of(k)}" for k in inst.K
        ),
        "primary_obj": inst.config["primary_obj"],
        "constraint_obj": inst.config["constraint_obj"],
        "objective_method": inst.config["objective_method"],
        "limit_on_constraint_obj": inst.config["limit_on_constraint_obj"],
        "augmentation_weight": (inst.config["augmentation_weight"]
                                if inst.config["objective_method"] == "augmented_eps"
                                else "(unused: lexicographic)"),
        "route_duration_min": route_duration_val,
        "total_wait_min": total_wait_val,
        "obj_value": obj_value,
        "best_bound": best_bound,
        "mip_gap": mip_gap,
        "runtime_s": _safe_attr(m, "Runtime"),
        "status": _safe_attr(m, "Status"),
        "T_max": inst.T_max, "C_max": inst.C_max,
        "e_min": inst.e_min, "Q_max": inst.Q_max,
        "big_m_mode": inst.big_m_mode,
        "M16_pdf": inst.M16_pdf, "M16_used": inst.M16,
        "M20": inst.M20,
        "M22_pdf_uniform": inst.M22_pdf,
        "M22_used_min": min(inst.M22_p.values()),
        "M22_used_max": max(inst.M22_p.values()),
        "M24": inst.M24, "M25": inst.M25, "U": inst.U,
        "shift_start_clock_min": off,
        "shift_duration_min": inst.config["shift_duration_min"],
        "break_vehicle_symmetry": inst.config["break_vehicle_symmetry"],
        "vehicle_symmetry_groups": (
            "; ".join(",".join(g) for g in inst.vehicle_symmetry_groups)
            if inst.vehicle_symmetry_groups else "none"
        ),
        "add_work_lb_cut": inst.config["add_work_lb_cut"],
        "tight_time_var_bounds": inst.config["tight_time_var_bounds"],
        "tight_route_activation": inst.config["tight_route_activation"],
        "use_indicator_constraints": inst.config["use_indicator_constraints"],
        "add_product_lb_cut": inst.config["add_product_lb_cut"],
        "add_pair_lb_cut": inst.config["add_pair_lb_cut"],
        "pair_lb_threshold_min": inst.config["pair_lb_threshold_min"],
        "pair_lb_cuts_emitted_count": len(inst.pair_lb_cuts),
        "add_wait_lb_cut": inst.config["add_wait_lb_cut"],
        "add_reverse_arc_cut": inst.config["add_reverse_arc_cut"],
        "add_endpoints_cut": inst.config["add_endpoints_cut"],
        "add_adjacency_cut": inst.config["add_adjacency_cut"],
        "mtz_type": inst.config["mtz_type"],
    }

    full = inst.config.get("write_full_var_sheets", True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame([summary]).to_excel(writer, sheet_name="summary", index=False)

        # Per-product Big-M values used in (22) — for transparency
        rows = [{"product": p,
                 "ready_relative": inst.e[p],
                 "s_unload": inst.s_unload[p],
                 "M22_pdf_uniform": inst.M22_pdf,
                 "M22_used": inst.M22_p[p]}
                for p in inst.P]
        pd.DataFrame(rows).to_excel(writer, sheet_name="big_m_per_product", index=False)

        # Vehicle symmetry groups detected (informational; constraint only added if break_vehicle_symmetry=true)
        sym_rows = []
        for gi, group in enumerate(inst.vehicle_symmetry_groups, start=1):
            for vid in group:
                sym_rows.append({"group": f"G{gi}", "vehicle_id": vid,
                                 "capacity_m2": inst.q_k[vid]})
        if not sym_rows:
            sym_rows = [{"group": "(none)", "vehicle_id": "", "capacity_m2": ""}]
        pd.DataFrame(sym_rows).to_excel(writer, sheet_name="symmetry_groups", index=False)

        # Active arcs
        rows = [{"i": i, "j": j, "k": k, "r": r,
                 "c_ij_min": inst.c[(i, j)]}
                for (i, j, k, r), v in x.items() if v.X > 0.5]
        pd.DataFrame(rows).to_excel(writer, sheet_name="x_used", index=False)

        # Assignments
        rows = [{"product": p, "k": k, "r": r,
                 "origin": inst.o[p], "destination": inst.d[p]}
                for (p, k, r), v in f.items() if v.X > 0.5]
        pd.DataFrame(rows).to_excel(writer, sheet_name="assignment_f", index=False)

        # Wait times
        rows = [{"product": p,
                 "ready_relative": inst.e[p],
                 "ready_clock": stamp(inst.e[p]),
                 "wait_min": w[p].X}
                for p in inst.P]
        pd.DataFrame(rows).to_excel(writer, sheet_name="wait_w", index=False)

        # Itinerary (used routes only)
        itin = []
        for k in inst.K:
            for r in inst.routes_of(k):
                used = any(x["h", j, k, r].X > 0.5 for j in inst.Nw)
                if not used:
                    continue
                cur = "h"
                visited = {"h"}
                order = 1
                itin.append({"k": k, "r": r, "order": order, "node": "h",
                             "ta": ta["h", k, r].X, "td": td["h", k, r].X,
                             "ta_clock": stamp(ta["h", k, r].X),
                             "td_clock": stamp(td["h", k, r].X),
                             "y_after": y["h", k, r].X})
                while True:
                    nxt = None
                    for j in inst.N:
                        if j != cur and (cur, j, k, r) in x and x[cur, j, k, r].X > 0.5:
                            nxt = j
                            break
                    if nxt is None:
                        break
                    order += 1
                    itin.append({"k": k, "r": r, "order": order, "node": nxt,
                                 "ta": ta[nxt, k, r].X, "td": td[nxt, k, r].X,
                                 "ta_clock": stamp(ta[nxt, k, r].X),
                                 "td_clock": stamp(td[nxt, k, r].X),
                                 "y_after": y[nxt, k, r].X})
                    if nxt == "h" or nxt in visited:
                        break
                    visited.add(nxt)
                    cur = nxt
        pd.DataFrame(itin).to_excel(writer, sheet_name="itinerary", index=False)

        # ---- route_timings sheet: one row per traversed arc ----
        # For each used (vehicle, route), walk the route in order and emit
        # one row per arc with depart-from-i / arrive-at-j clock stamps,
        # travel time, and the load / unload service times happening at
        # the endpoints. A TOTAL row per route summarises travel + service.
        timing_rows = []
        for k in inst.K:
            for r in inst.routes_of(k):
                if not any(x["h", j, k, r].X > 0.5 for j in inst.Nw):
                    continue
                seq = ["h"]
                cur = "h"
                visited = {"h"}
                while True:
                    nxt = None
                    for j in inst.N:
                        if j != cur and (cur, j, k, r) in x and x[cur, j, k, r].X > 0.5:
                            nxt = j
                            break
                    if nxt is None:
                        break
                    seq.append(nxt)
                    if nxt == "h" or nxt in visited:
                        break
                    visited.add(nxt)
                    cur = nxt
                travel_sum = 0.0
                service_sum = 0.0
                for leg, (i_node, j_node) in enumerate(
                        zip(seq[:-1], seq[1:]), start=1):
                    depart_i = td[i_node, k, r].X if (i_node, k, r) in td else 0.0
                    arrive_j = ta[j_node, k, r].X if (j_node, k, r) in ta else 0.0
                    travel = inst.c.get((i_node, j_node), 0.0)
                    load_at_i = sum(
                        inst.s_load[p] for p in inst.P
                        if inst.o[p] == i_node
                        and (p, k, r) in f and f[p, k, r].X > 0.5
                    )
                    unload_at_j = sum(
                        inst.s_unload[p] for p in inst.P
                        if inst.d[p] == j_node
                        and (p, k, r) in f and f[p, k, r].X > 0.5
                    )
                    travel_sum += travel
                    service_sum += load_at_i + unload_at_j
                    timing_rows.append({
                        "vehicle": k,
                        "route": r,
                        "leg": leg,
                        "from": i_node,
                        "to": j_node,
                        "depart_min": round(depart_i, 3),
                        "depart_clock": stamp(depart_i),
                        "arrive_min": round(arrive_j, 3),
                        "arrive_clock": stamp(arrive_j),
                        "travel_min": round(travel, 3),
                        "load_min_at_from": round(load_at_i, 3),
                        "unload_min_at_to": round(unload_at_j, 3),
                    })
                # Per-route totals row
                timing_rows.append({
                    "vehicle": k,
                    "route": r,
                    "leg": "TOTAL",
                    "from": "—",
                    "to": "—",
                    "depart_min": "",
                    "depart_clock": "",
                    "arrive_min": "",
                    "arrive_clock": "",
                    "travel_min": round(travel_sum, 3),
                    "load_min_at_from": "",
                    "unload_min_at_to": "",
                })
                timing_rows.append({
                    "vehicle": k,
                    "route": r,
                    "leg": "SERVICE",
                    "from": "—",
                    "to": "—",
                    "depart_min": "",
                    "depart_clock": "",
                    "arrive_min": "",
                    "arrive_clock": "",
                    "travel_min": round(service_sum, 3),
                    "load_min_at_from": "",
                    "unload_min_at_to": "",
                })
        if timing_rows:
            pd.DataFrame(timing_rows).to_excel(
                writer, sheet_name="route_timings", index=False
            )

        if full:
            rows = [{"node": j, "k": k, "r": r,
                     "ta": ta[j, k, r].X, "ta_clock": stamp(ta[j, k, r].X),
                     "td": td[j, k, r].X, "td_clock": stamp(td[j, k, r].X),
                     "y_after": y[j, k, r].X}
                    for (k, r) in inst.KR_pairs for j in inst.N]
            pd.DataFrame(rows).to_excel(writer, sheet_name="node_times", index=False)

            rows = [{"node": j, "k": k, "r": r,
                     "ts": ts[j, k, r].X, "ts_clock": stamp(ts[j, k, r].X),
                     "u": u[j, k, r].X, "delta": delta[j, k, r].X}
                    for (k, r) in inst.KR_pairs for j in inst.Nw]
            pd.DataFrame(rows).to_excel(writer, sheet_name="service_u_delta", index=False)

    log.info("Results written to %s", output_path)
    return summary


# =============================================================================
# Post-solve: verify + visualize
# =============================================================================
def postprocess(m, vars_, inst, cfg, result_xlsx, run_dir, timestamp: str = ""):
    """Run verifier and visualizer after a successful solve.

    All artefacts (validation.txt, gantt.png, routes.png, wait.png) are
    written to `run_dir`. Honours `auto_verify`, `verify_on_fail`, and
    `auto_visualize` config flags. On verification failure with
    `verify_on_fail = raise`, raises RuntimeError AFTER writing the
    validation report (so failure context survives).
    """
    verify_failed = False

    if cfg["auto_verify"]:
        try:
            from verify import (
                extract_solution, validate_solution, format_report, passed,
            )
        except ImportError as exc:
            print(f"[verify] could not import verify.py: {exc}")
            log.error("verify.py import failed: %s", exc)
            return

        sol = extract_solution(m, vars_, inst)
        findings = validate_solution(sol, inst)
        report_text = format_report(findings)

        report_path = run_dir / "validation.txt"
        report_path.write_text(report_text, encoding="utf-8")
        log.info("Verification report -> %s", report_path)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(result_xlsx)
            if "validation" in wb.sheetnames:
                del wb["validation"]
            ws = wb.create_sheet("validation")
            ws.append(["category", "test_id", "passed", "detail"])
            for f in findings:
                ws.append([f.category, f.test_id, f.passed, f.detail])
            wb.save(result_xlsx)
        except Exception as exc:
            log.warning("could not append validation sheet to xlsx: %s", exc)

        if not passed(findings):
            verify_failed = True
            print()
            print("[verify] VERIFICATION FAILED — see report:")
            print(report_text)
        else:
            n = len(findings)
            print(f"[verify] {n}/{n} checks passed")
            log.info("verification: all %d checks passed", n)
    else:
        log.info("verification skipped (auto_verify=false)")

    if cfg["auto_visualize"]:
        try:
            from verify import extract_solution
            from visualize import render_all
            sol_for_plots = extract_solution(m, vars_, inst)
            paths = render_all(sol_for_plots, inst, run_dir, suffix=timestamp)
            for kind, p in paths.items():
                print(f"[visualize] {kind} -> {p}")
            log.info("visualizations rendered: %s", list(paths.values()))
        except ImportError as exc:
            print(f"[visualize] could not import visualize.py: {exc}")
            log.error("visualize.py import failed: %s", exc)
        except Exception as exc:
            print(f"[visualize] FAILED: {exc}")
            log.error("visualization failed: %s", exc, exc_info=True)
    else:
        log.info("visualization skipped (auto_visualize=false)")

    if verify_failed and cfg["verify_on_fail"] == "raise":
        raise RuntimeError(
            "Solution failed verification (verify_on_fail=raise). "
            "See report above and in the .validation.txt file."
        )


# =============================================================================
# Multi-objective ε-constraint sweep
# =============================================================================
def _get_constraint_value(summary: dict, constraint_obj: str) -> float:
    """Pull the secondary-objective value from a per-iteration summary."""
    if constraint_obj == "wait_time":
        return float(summary["total_wait_min"])
    if constraint_obj == "route_duration":
        return float(summary["route_duration_min"])
    raise ValueError(f"unknown constraint_obj: {constraint_obj}")


def run_multi_objective_sweep(cfg, inst, main_dir: Path,
                                timestamp: str = "") -> list:
    """Adaptive ε-constraint sweep generating a Pareto frontier.

    Strategy: start with the user's initial limit_on_constraint_obj (typically
    a large value making the constraint non-binding); each subsequent iteration
    tightens ε to `previous_secondary_achieved - 1.0`. The sweep continues
    indefinitely until the model becomes infeasible — at that point all
    Pareto-optimal solutions have been enumerated.

    Per-iteration outputs go into `main_dir / "eps_<value>" /`. The aggregate
    `pareto_summary.xlsx` and `pareto_frontier.png` land in `main_dir`.
    """
    constraint_obj = cfg["constraint_obj"]
    primary_obj = cfg["primary_obj"]
    initial_limit = float(cfg["limit_on_constraint_obj"])
    main_dir.mkdir(parents=True, exist_ok=True)
    aggregate_rows: list = []
    current_limit = initial_limit

    print(f"[multiobj] start sweep with limit_on_constraint_obj = {initial_limit}")
    print(f"[multiobj] primary = {primary_obj}, constraint = {constraint_obj}")
    print(f"[multiobj] outputs -> {main_dir}")
    print()

    sub_cfg = dict(cfg)
    sub_cfg["verify_on_fail"] = "warn"  # never abort sweep on validation failure

    it = 0
    while True:
        sub_label = f"eps_{current_limit:.2f}"
        sub_dir = main_dir / sub_label
        sub_dir.mkdir(parents=True, exist_ok=True)
        sfx = f"_{timestamp}" if timestamp else ""
        gurobi_log = sub_dir / f"gurobi{sfx}.log"
        result_xlsx = sub_dir / f"result{sfx}.xlsx"

        print(f"[multiobj] --- iteration {it}: eps = {current_limit:.2f} ---")
        log.info("Iteration %d: building model with limit=%.4f", it, current_limit)

        m, vars_ = build_model(
            inst,
            primary=primary_obj,
            constraint=constraint_obj,
            limit=current_limit,
            weight=cfg["augmentation_weight"],
            method=cfg["objective_method"],
        )
        # Only request IIS for the very first iteration (the unconstrained
        # case). Infeasibility there indicates a genuine data problem.
        # Tightening iterations that hit infeasibility are simply the
        # expected end of the Pareto sweep; running IIS there would
        # produce noise tied to the user-imposed eps constraint.
        solve_model(
            m, cfg=cfg, gurobi_log=gurobi_log,
            compute_iis_on_infeasible=(it == 0),
        )

        # Detect infeasibility
        if m.SolCount == 0:
            print(f"[multiobj] iter {it}: INFEASIBLE at eps={current_limit:.2f} — stopping sweep")
            log.info("Iteration %d infeasible; terminating sweep", it)
            aggregate_rows.append({
                "iter": it,
                "epsilon_used": current_limit,
                "status": "infeasible",
                "subfolder": sub_label,
            })
            break

        summary = write_results(m, vars_, inst, output_path=result_xlsx)
        if summary is None:
            print(f"[multiobj] iter {it}: no solution captured; stopping sweep")
            break

        # Run verifier + visualizations inside the iteration's subfolder
        try:
            postprocess(m, vars_, inst, sub_cfg, result_xlsx, sub_dir,
                        timestamp=timestamp)
        except Exception as exc:
            log.warning("postprocess raised on iteration %d: %s", it, exc)
            print(f"[multiobj] iter {it}: postprocess warning: {exc}")

        achieved_secondary = _get_constraint_value(summary, constraint_obj)
        row = {
            "iter": it,
            "epsilon_used": current_limit,
            "route_duration_min": summary["route_duration_min"],
            "total_wait_min": summary["total_wait_min"],
            "obj_value": summary["obj_value"],
            "mip_gap": summary["mip_gap"],
            "runtime_s": summary["runtime_s"],
            "status": summary["status"],
            "subfolder": sub_label,
        }
        aggregate_rows.append(row)

        print(
            f"[multiobj] iter {it}: route_duration={summary['route_duration_min']:.2f}  "
            f"total_wait={summary['total_wait_min']:.2f}  "
            f"runtime={summary['runtime_s']:.2f}s"
        )

        # Tighten epsilon for next iteration. Step size controlled by
        # cfg["eps_step"] (default 1.0). Smaller step = denser Pareto sweep
        # but more iterations.
        step = float(cfg.get("eps_step", 1.0))
        new_limit = achieved_secondary - step
        if new_limit < 0:
            print(f"[multiobj] next eps = {new_limit:.2f} < 0 — stopping sweep")
            break
        current_limit = new_limit
        it += 1

    # ---- Aggregate spreadsheet ----
    sfx = f"_{timestamp}" if timestamp else ""
    sheet_path = main_dir / f"pareto_summary{sfx}.xlsx"
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "pareto_summary"
        cols = ["iter", "epsilon_used", "route_duration_min", "total_wait_min",
                "obj_value", "mip_gap", "runtime_s", "status", "subfolder"]
        ws.append(cols)
        for row in aggregate_rows:
            ws.append([row.get(c) for c in cols])
        # echo config as a second sheet for reproducibility
        ws_cfg = wb.create_sheet("config_used")
        ws_cfg.append(["parameter", "value"])
        for k, v in sorted(cfg.items()):
            ws_cfg.append([k, str(v)])
        wb.save(sheet_path)
        print(f"[multiobj] aggregate -> {sheet_path}")
    except Exception as exc:
        log.error("could not write pareto_summary.xlsx: %s", exc)
        print(f"[multiobj] failed to write summary xlsx: {exc}")

    # ---- Pareto frontier plot ----
    plot_path = main_dir / f"pareto_frontier{sfx}.png"
    try:
        from visualize import plot_pareto_frontier
        plot_pareto_frontier(
            aggregate_rows, plot_path,
            x_label=(f"{primary_obj} (min)"),
            y_label=(f"{constraint_obj} (min)"),
            title=f"Pareto frontier — {cfg['product_set_id']} (|P|={len(inst.P)})",
        )
        print(f"[multiobj] frontier -> {plot_path}")
    except Exception as exc:
        log.error("could not generate pareto_frontier.png: %s", exc)
        print(f"[multiobj] failed to plot frontier: {exc}")

    return aggregate_rows


# =============================================================================
# Main
# =============================================================================
def main() -> int:
    parser = argparse.ArgumentParser(description="Internal Logistics MIP runner")
    parser.add_argument("--inputs", type=Path, default=Path("inputs"))
    parser.add_argument("--config", type=Path, default=None,
                        help="Path to config.xlsx (default: <inputs>/config.xlsx)")
    parser.add_argument("--output-dir", type=Path, default=None,
                        help="Override output_dir from config")
    parser.add_argument("--verbose", action="store_true",
                        help="Also stream Python log records to the console")
    args = parser.parse_args()

    config_path = args.config or (args.inputs / "config.xlsx")
    cfg = load_config(config_path)
    out_dir = (Path(args.output_dir) if args.output_dir
               else Path(cfg["output_dir"]))
    out_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
    # One output folder per run; its name depends on the run mode so that
    # multi-objective runs land directly in their multiobj_* folder without
    # also creating a redundant run_* sibling.
    if cfg["run_mode"] == "single_objective":
        label = (f"{cfg['output_prefix']}_{cfg['product_set_id']}"
                 f"_{cfg['primary_obj']}_{timestamp}")
    else:  # multi_objective
        label = (f"multiobj_{cfg['product_set_id']}"
                 f"_{cfg['primary_obj']}_{timestamp}")
    run_dir = out_dir / label
    run_dir.mkdir(parents=True, exist_ok=True)
    # Every file inside the run folder carries the same timestamp suffix
    # as the folder itself, so artefacts from different runs (which may
    # share the same folder during testing) stay distinguishable when
    # files are copied out for archiving or sharing.
    log_path = run_dir / f"run_{timestamp}.log"
    gurobi_log = run_dir / f"gurobi_{timestamp}.log"
    result_xlsx = run_dir / f"result_{timestamp}.xlsx"

    configure_logging(log_file=log_path, verbose=args.verbose)

    print(f"[run] {label}")
    print(f"[run] config  : {config_path}")
    print(f"[run] inputs  : {args.inputs}")
    print(f"[run] output  : {out_dir}")
    print(f"[run] script log -> {log_path}")
    if cfg["run_mode"] == "single_objective":
        print(f"[run] gurobi log -> {gurobi_log}")
    print()

    log.info("Run label : %s", label)
    log.info("Config    : %s", config_path)
    log.info("Inputs    : %s", args.inputs)
    log.info("Output    : %s", out_dir)

    try:
        inst = load_instance(args.inputs, cfg)
    except DataConsistencyError as exc:
        print(f"[run] DATA VALIDATION FAILED: {exc}")
        log.error("DATA VALIDATION FAILED: %s", exc)
        return 2

    if cfg["run_mode"] == "single_objective":
        m, vars_ = build_model(
            inst,
            primary=cfg["primary_obj"],
            constraint=cfg["constraint_obj"],
            limit=cfg["limit_on_constraint_obj"],
            weight=cfg["augmentation_weight"],
            method=cfg["objective_method"],
        )
        solve_model(m, cfg=cfg, gurobi_log=gurobi_log)
        summary = write_results(m, vars_, inst, output_path=result_xlsx)
        if summary:
            print()
            print(
                f"[run] DONE  route_duration={summary['route_duration_min']:.2f}  "
                f"total_wait={summary['total_wait_min']:.2f}  "
                f"obj={summary['obj_value']:.4f}  "
                f"runtime={summary['runtime_s']:.2f}s"
            )
            print(f"[run] results -> {result_xlsx}")
            log.info(
                "DONE - route_duration=%.2f total_wait=%.2f obj=%.4f runtime=%.2fs",
                summary["route_duration_min"],
                summary["total_wait_min"],
                summary["obj_value"],
                summary["runtime_s"],
            )
            postprocess(m, vars_, inst, cfg, result_xlsx, run_dir,
                        timestamp=timestamp)
        else:
            print("[run] no feasible solution found")
        return 0

    # multi_objective mode: adaptive epsilon-constraint sweep.
    # The already-created `run_dir` is the multiobj folder (its label was set
    # accordingly above), so we reuse it as `main_dir` for the sweep — no
    # redundant `run_*` sibling is created.
    print(f"[run] multi_objective mode -> {run_dir}")
    log.info("Multi-objective sweep folder: %s", run_dir)
    run_multi_objective_sweep(cfg, inst, run_dir, timestamp=timestamp)
    return 0


if __name__ == "__main__":
    sys.exit(main())
